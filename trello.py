import requests
import pandas as pd
import re
from collections import Counter
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
import json

# Load configuration from config.json
with open('config.json', 'r') as config_file:
    config = json.load(config_file)

# Your Trello API key, token, and board ID from config.json
API_KEY = config['API_KEY']
TOKEN = config['TOKEN']
BOARD_ID = config['BOARD_ID']
APPRENTICE_NAME = config['NAME']
MONTH = config['MONTH']
YEAR = config['YEAR']

# Function to fetch cards from a board
def get_cards(board_id, api_key, token):
    url = f'https://api.trello.com/1/boards/{board_id}/cards'
    query = {
        'key': api_key,
        'token': token,
    }
    response = requests.get(url, params=query)
    response.raise_for_status()  # Raise an error if the request fails
    return response.json()

# Function to fetch comments for a card
def get_comments(card_id, api_key, token):
    url = f'https://api.trello.com/1/cards/{card_id}/actions'
    query = {
        'key': api_key,
        'token': token,
        'filter': 'commentCard'
    }
    response = requests.get(url, params=query)
    response.raise_for_status()  # Raise an error if the request fails
    return response.json()

# Function to extract the date from a comment text
def extract_date(comment_text):
    date_patterns = [
        r'\b\d{1,2}/\d{1,2}/\d{4}\b',  # D/M/YYYY or DD/M/YYYY or D/MM/YYYY or DD/MM/YYYY
        r'\b\d{1,2}/\d{1,2}/\d{2}\b',   # D/M/YY or DD/M/YY or D/MM/YY or DD/MM/YY
        r'\b\d{4}-\d{2}-\d{2}\b',       # YYYY-MM-DD
        r'\b\d{2}-\d{2}-\d{4}\b',       # DD-MM-YYYY
        r'\b\d{2}-\d{2}-\d{2}\b',       # DD-MM-YY
        r'\b\d{1,2}/\d{1,2}\b'          # D/M or DD/M or D/MM or DD/MM (no year)
    ]
    
    for pattern in date_patterns:
        match = re.search(pattern, comment_text)
        if match:
            date_str = match.group(0)
            for fmt in ['%d/%m/%Y', '%d/%m/%y', '%d-%m-%Y', '%d-%m-%y', '%Y-%m-%d']:
                try:
                    date_obj = datetime.strptime(date_str, fmt)
                    formatted_date = date_obj.strftime('%d/%m/%Y')
                    return formatted_date, re.sub(pattern, '', comment_text).strip()
                except ValueError:
                    continue
            if len(date_str.split('/')) == 2 or len(date_str.split('-')) == 2:
                date_str_with_year = f"{date_str}/24" if '/' in date_str else f"{date_str}-24"
                for fmt in ['%d/%m/%y', '%d-%m-%y']:
                    try:
                        date_obj = datetime.strptime(date_str_with_year, fmt)
                        formatted_date = date_obj.strftime('%d/%m/%Y')
                        return formatted_date, re.sub(pattern, '', comment_text).strip()
                    except ValueError:
                        continue
    return None, comment_text

# Function to extract duration from the text
def extract_duration(text):
    match = re.search(r'\[(\d+)\]', text)
    if match:
        duration = match.group(1)
        text = re.sub(r'\s*\[\d+\]\s*', ' ', text).strip()
        return int(duration), text
    return None, text

# Function to determine knowledge, skill, or behaviour from the text
def determine_ksb(text):
    if '[K]' in text:
        return 'Knowledge', re.sub(r'\s*\[K\]\s*', ' ', text).strip()
    elif '[S]' in text:
        return 'Skill', re.sub(r'\s*\[S\]\s*', ' ', text).strip()
    elif '[B]' in text:
        return 'Behaviour', re.sub(r'\s*\[B\]\s*', ' ', text).strip()
    else:
        return None, text  # Return None if no marker is found

# Dropdown options for the new columns
learning_activity_options = [
    "Online Learning",
    "TCG Set Tasks",
    "Portfolio Work (non-admin)",
    "Coaching / Mentoring",
    "Event / Meetup / Conference",
    "Shadowing",
    "CPD",
    "Workplace Training",
    "Research",
    "Peer to Peer Support",
    "Workplace Reviews"
]

ksb_contribution_options = ["Yes", "No"]

knowledge_skill_behaviour_options = ["Knowledge", "Skills", "Behaviour"]

# Fetch cards
cards = get_cards(BOARD_ID, API_KEY, TOKEN)

# Prepare data for the DataFrame
data = []
all_dates = []

for card in cards:
    card_id = card['id']
    card_name = card['name']
    card_desc = card['desc']
    card_url = card['shortUrl']
    
    comments = get_comments(card_id, API_KEY, TOKEN)
    
    if not comments:
        continue  # Skip cards with no comments
    else:
        for comment in comments:
            comment_text = comment['data']['text']
            if '(X)' in comment_text or '[X]' in comment_text:
                continue  # Skip rows with (X) or [X] in the comments
            extracted_date, cleaned_desc = extract_date(comment_text)
            if extracted_date:
                all_dates.append(extracted_date)
                comment_date_obj = datetime.strptime(extracted_date, '%d/%m/%Y')
                if comment_date_obj.month != int(MONTH) or comment_date_obj.year != int(YEAR):
                    continue  # Skip comments not in the specified month and year
            
            # Extract duration and KSB from both card description and comment
            duration, updated_desc = extract_duration(card_desc)
            duration_from_comment, cleaned_comment_desc = extract_duration(cleaned_desc)
            if duration_from_comment is not None:
                duration = duration_from_comment
            
            ksb_type, combined_desc = determine_ksb(f"{updated_desc}\n{cleaned_comment_desc}".strip())
            if ksb_type is None:
                ksb_type = 'Knowledge'  # Default to 'Knowledge' if no marker is found
            
            data.append({
                'Date': extracted_date,
                'Type of Learning Activity': 'Online Learning',
                'Details of Learning Activity': card_name,
                'What did you learn from this and how will you use what you have learnt?': combined_desc,
                'Does it contribute to KSBs?': 'Yes',
                'Time (hrs)': duration,
                'Knowledge, Skill, Behaviour': ksb_type
            })

# Create a DataFrame
df = pd.DataFrame(data)

# Ensure Time (hrs) column is numeric
df['Time (hrs)'] = pd.to_numeric(df['Time (hrs)'], errors='coerce')

# Construct the output file name with the month name
month_name = datetime.strptime(MONTH, '%m').strftime('%B')
output_file = f"{APPRENTICE_NAME} OTJ Log {month_name} {YEAR}.xlsx"

# Export to Excel
df.to_excel(output_file, index=False)

# Load the workbook and worksheet
wb = load_workbook(output_file)
ws = wb.active

# Adjust column widths to a reasonable size
reasonable_widths = {
    'A': 22,  # Date
    'B': 25,  # Type of Learning Activity
    'C': 70,  # Details of Learning Activity
    'D': 100,  # What did you learn from this and how will you use what you have learnt?
    'E': 25,  # Does it contribute to KSBs?
    'F': 10,  # Time (hrs)
    'G': 40   # Knowledge, Skill, Behaviour
}

for col_letter, width in reasonable_widths.items():
    ws.column_dimensions[col_letter].width = width

# Align all column headings to the left
for cell in ws["6:6"]:
    cell.alignment = Alignment(horizontal='left')

# Insert new rows at the top
ws.insert_rows(1, 5)

# Add Apprentice Name, Month, Year, and OFF THE JOB - MONTHLY EVIDENCE
dark_grey_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

ws['A2'] = 'Apprentice Name:'
ws['A3'] = 'Month:'
ws['A4'] = 'Year:'
for cell in ['A2', 'A3', 'A4']:
    ws[cell].font = Font(bold=True)
    ws[cell].fill = dark_grey_fill
for cell in ['B2', 'B3', 'B4']:
    ws[cell].fill = light_grey_fill
ws['B2'] = APPRENTICE_NAME  # Populate cell B2 with the apprentice name from config.json
ws['B3'] = month_name
ws['B4'] = YEAR
ws.merge_cells('C2:D4')
ws['C2'] = 'OFF THE JOB - MONTHLY EVIDENCE'
ws['C2'].font = Font(size=24, bold=True)
ws['C2'].alignment = Alignment(horizontal='center', vertical='center')

# Create data validation objects
learning_activity_dv = DataValidation(
    type="list",
    formula1=f'"{",".join(learning_activity_options)}"',
    showDropDown=True
)

ksb_contribution_dv = DataValidation(
    type="list",
    formula1=f'"{",".join(ksb_contribution_options)}"',
    showDropDown=True
)

knowledge_skill_behaviour_dv = DataValidation(
    type="list",
    formula1=f'"{",".join(knowledge_skill_behaviour_options)}"',
    showDropDown=True
)

# Add the data validation to the Type of Learning Activity column
learning_activity_col = get_column_letter(df.columns.get_loc('Type of Learning Activity') + 1)
for row in range(7, len(df) + 7):
    cell = f"{learning_activity_col}{row}"
    ws[cell].value = "Online Learning"
    ws.add_data_validation(learning_activity_dv)
    learning_activity_dv.add(ws[cell])

# Add the data validation to the Does it contribute to KSBs? column
ksb_contribution_col = get_column_letter(df.columns.get_loc('Does it contribute to KSBs?') + 1)
for row in range(7, len(df) + 7):
    cell = f"{ksb_contribution_col}{row}"
    ws.add_data_validation(ksb_contribution_dv)
    ksb_contribution_dv.add(ws[cell])
    ws[cell].alignment = Alignment(horizontal='left')

# Add the data validation to the Knowledge, Skill, Behaviour column
knowledge_skill_behaviour_col = get_column_letter(df.columns.get_loc('Knowledge, Skill, Behaviour') + 1)
for row in range(7, len(df) + 7):
    cell = f"{knowledge_skill_behaviour_col}{row}"
    ws.add_data_validation(knowledge_skill_behaviour_dv)
    knowledge_skill_behaviour_dv.add(ws[cell])

# Calculate total off-the-job hours
time_hrs_col = get_column_letter(df.columns.get_loc('Time (hrs)') + 1)
total_hours_cell = f"{time_hrs_col}{len(df) + 7}"
ws[total_hours_cell] = f"=SUM({time_hrs_col}7:{time_hrs_col}{len(df) + 6})"
ws[f"{ksb_contribution_col}{len(df) + 7}"] = "TOTAL OFF-THE-JOB HOURS"

# Add declaration
declaration_text = """
DECLARATION
· The training listed above has been undertaken within my normal working hours
· The training is directly relevant to, and provided new knowledge, skills or behaviours required to, achieve my apprenticeship.
"""
declaration_start_row = len(df) + 9
ws.merge_cells(start_row=declaration_start_row, start_column=1, end_row=declaration_start_row + 4, end_column=4)
declaration_cell = ws.cell(row=declaration_start_row, column=1)
declaration_cell.value = declaration_text
declaration_cell.alignment = Alignment(wrap_text=True)

# Save the workbook
wb.save(output_file)

print(f'Data successfully exported to {output_file}')
